"""Paginated reader for original uploaded CSV and Excel files.

Both readers stream the source file in bounded batches so a page request
never materializes the full file in memory — important for large local test
files where the web process has a fixed memory budget (e.g. the Docker
web container).
"""

from __future__ import annotations

import math
import warnings
from datetime import date, datetime
from decimal import Decimal
from typing import Any

import openpyxl
import pyarrow as pa
import pyarrow.csv as pcsv

from nltoregex.cache_utils import get_cached_upload_page, set_cached_upload_page

from .excel_reader import resolve_excel_sheet_name


def _json_safe(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


def _batch_to_records(batch: pa.RecordBatch) -> list[dict[str, Any]]:
    columns = batch.to_pydict()
    if not columns:
        return []

    column_names = list(columns.keys())
    row_count = batch.num_rows
    return [
        {
            column: _json_safe(columns[column][row_index])
            for column in column_names
        }
        for row_index in range(row_count)
    ]


def _paginate_csv(file_path: str, page: int, page_size: int) -> dict[str, Any]:
    """Stream the CSV in Arrow batches, keeping only the requested page in memory."""
    reader = pcsv.open_csv(file_path)
    columns = reader.schema.names

    offset = (page - 1) * page_size
    collected_rows: list[dict[str, Any]] = []
    total_rows = 0
    skipped_rows = 0

    for batch in reader:
        batch_row_count = batch.num_rows
        total_rows += batch_row_count

        if len(collected_rows) >= page_size:
            continue

        if skipped_rows + batch_row_count <= offset:
            skipped_rows += batch_row_count
            continue

        start_in_batch = max(0, offset - skipped_rows)
        remaining = page_size - len(collected_rows)
        slice_length = min(batch_row_count - start_in_batch, remaining)
        sliced_batch = batch.slice(start_in_batch, slice_length)
        collected_rows.extend(_batch_to_records(sliced_batch))
        skipped_rows += batch_row_count

    if total_rows == 0:
        return {"columns": columns, "rows": [], "total_rows": 0, "total_pages": 0}

    return {
        "columns": columns,
        "rows": collected_rows,
        "total_rows": total_rows,
        "total_pages": math.ceil(total_rows / page_size),
    }


def _paginate_excel(file_path: str, page: int, page_size: int) -> dict[str, Any]:
    """Stream the worksheet row-by-row via openpyxl read-only mode."""
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    try:
        sheet_name = resolve_excel_sheet_name(file_path, workbook.sheetnames)
        sheet = workbook[sheet_name]
        rows_iter = sheet.iter_rows(values_only=True)

        try:
            header_row = next(rows_iter)
        except StopIteration:
            return {"columns": [], "rows": [], "total_rows": 0, "total_pages": 0}

        columns = [str(cell).strip() if cell is not None else "" for cell in header_row]

        offset = (page - 1) * page_size
        collected_rows: list[dict[str, Any]] = []
        total_rows = 0

        for row in rows_iter:
            if offset <= total_rows < offset + page_size:
                collected_rows.append(
                    {
                        column: _json_safe(value)
                        for column, value in zip(columns, row)
                    }
                )
            total_rows += 1
    finally:
        workbook.close()

    if total_rows == 0:
        return {"columns": columns, "rows": [], "total_rows": 0, "total_pages": 0}

    return {
        "columns": columns,
        "rows": collected_rows,
        "total_rows": total_rows,
        "total_pages": math.ceil(total_rows / page_size),
    }


def paginate_uploaded_file(
    file_path: str,
    page: int,
    page_size: int,
    uploaded_file_id: str | None = None,
) -> dict[str, Any]:
    """Return one page of rows from an uploaded CSV or Excel file.

    Streams the source file rather than loading it fully into memory, and
    caches the page in Redis (keyed by upload id) so repeated views — e.g.
    reopening a large file from history — don't re-scan the whole file.
    """
    if uploaded_file_id:
        cached_page = get_cached_upload_page(uploaded_file_id, page, page_size)
        if cached_page is not None:
            return cached_page

    lower = file_path.lower()
    if lower.endswith(".csv"):
        page_data = _paginate_csv(file_path, page, page_size)
    elif lower.endswith(".xlsx"):
        page_data = _paginate_excel(file_path, page, page_size)
    else:
        raise ValueError(f"Unsupported file type for preview: {file_path}")

    if uploaded_file_id:
        set_cached_upload_page(uploaded_file_id, page, page_size, page_data)

    return page_data
