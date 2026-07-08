"""Read Excel workbooks directly from the XLSX zip/XML structure.

openpyxl fails on some files (for example strict OOXML exports), so this module
falls back to parsing workbook.xml and worksheet XML when needed.
"""
import csv
import logging
import warnings
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

import openpyxl
from django.conf import settings
from openpyxl.utils.cell import column_index_from_string

logger = logging.getLogger(__name__)


class ExcelReadError(ValueError):
    pass


def _xml_local_tag(element: ET.Element) -> str:
    """Return the XML tag name without its namespace prefix."""
    return element.tag.rsplit("}", 1)[-1]


def _relationship_id(element: ET.Element) -> str | None:
    for key, value in element.attrib.items():
        if key == "r:id" or key.endswith("}id"):
            return value
    return None


def _sheet_names_from_zip(file_path: str) -> list[str]:
    """Read sheet names from workbook.xml without relying on openpyxl."""
    with zipfile.ZipFile(file_path) as archive:
        with archive.open("xl/workbook.xml") as handle:
            root = ET.parse(handle).getroot()

    names: list[str] = []
    for element in root.iter():
        if _xml_local_tag(element) == "sheet":
            name = element.get("name")
            if name:
                names.append(name)
    return names


def _worksheet_paths_by_name(file_path: str) -> dict[str, str]:
    with zipfile.ZipFile(file_path) as archive:
        with archive.open("xl/workbook.xml") as handle:
            workbook_root = ET.parse(handle).getroot()
        with archive.open("xl/_rels/workbook.xml.rels") as handle:
            rels_root = ET.parse(handle).getroot()

    targets_by_id: dict[str, str] = {}
    for relationship in rels_root:
        if _xml_local_tag(relationship) != "Relationship":
            continue
        rel_id = relationship.get("Id")
        target = relationship.get("Target")
        rel_type = relationship.get("Type", "")
        if rel_id and target and rel_type.endswith("/worksheet"):
            targets_by_id[rel_id] = f"xl/{target.lstrip('/')}"

    worksheet_paths: dict[str, str] = {}
    for element in workbook_root.iter():
        if _xml_local_tag(element) != "sheet":
            continue
        sheet_name = element.get("name")
        rel_id = _relationship_id(element)
        if sheet_name and rel_id and rel_id in targets_by_id:
            worksheet_paths[sheet_name] = targets_by_id[rel_id]

    if not worksheet_paths:
        # Some malformed files omit workbook relationships; fall back to sheet1.xml.
        worksheet_files = sorted(
            name
            for name in zipfile.ZipFile(file_path).namelist()
            if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")
        )
        sheet_names = _sheet_names_from_zip(file_path)
        for index, worksheet_path in enumerate(worksheet_files):
            sheet_name = sheet_names[index] if index < len(sheet_names) else f"Sheet{index + 1}"
            worksheet_paths[sheet_name] = worksheet_path

    return worksheet_paths


def _shared_strings_from_zip(archive: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []

    with archive.open("xl/sharedStrings.xml") as handle:
        root = ET.parse(handle).getroot()

    shared_strings: list[str] = []
    for element in root.iter():
        if _xml_local_tag(element) != "si":
            continue
        parts = [node.text or "" for node in element.iter() if _xml_local_tag(node) == "t"]
        shared_strings.append("".join(parts))
    return shared_strings


def _cell_value(cell: ET.Element, shared_strings: list[str]) -> str:
    value_element = next((node for node in cell if _xml_local_tag(node) == "v"), None)
    if value_element is None or value_element.text is None:
        inline = next((node for node in cell.iter() if _xml_local_tag(node) == "t"), None)
        return inline.text.strip() if inline is not None and inline.text else ""

    raw_value = value_element.text
    if cell.get("t") == "s":
        # Shared-string cells store an index rather than the literal value.
        return shared_strings[int(raw_value)].strip()
    return raw_value.strip()


def _column_index_from_cell_ref(cell_ref: str) -> int:
    return column_index_from_string("".join(character for character in cell_ref if character.isalpha()))


def _row_values(row_element: ET.Element, shared_strings: list[str]) -> list[str]:
    cells: list[tuple[int, str]] = []
    for cell in row_element:
        if _xml_local_tag(cell) != "c":
            continue
        cell_ref = cell.get("r")
        if not cell_ref:
            continue
        cells.append(
            (
                _column_index_from_cell_ref(cell_ref),
                _cell_value(cell, shared_strings),
            )
        )

    if not cells:
        return []

    cells.sort(key=lambda item: item[0])
    max_column = cells[-1][0]
    values = [""] * max_column
    for column_index, value in cells:
        values[column_index - 1] = value
    return values


def _worksheet_rows_from_xml(
    archive: zipfile.ZipFile,
    worksheet_path: str,
    shared_strings: list[str],
) -> list[list[str]]:
    with archive.open(worksheet_path) as handle:
        root = ET.parse(handle).getroot()

    rows: list[list[str]] = []
    for row_element in root.iter():
        if _xml_local_tag(row_element) != "row":
            continue
        row_values = _row_values(row_element, shared_strings)
        if row_values:
            rows.append(row_values)

    if not rows:
        raise ExcelReadError("Excel worksheet has no data rows")

    max_width = max(len(row) for row in rows)
    return [row + [""] * (max_width - len(row)) for row in rows]


def _worksheet_path_for_file(file_path: str) -> str:
    sheet_name = resolve_excel_sheet_name(file_path)
    worksheet_paths = _worksheet_paths_by_name(file_path)
    worksheet_path = worksheet_paths.get(sheet_name)
    if worksheet_path is None:
        raise ExcelReadError(
            f"Worksheet '{sheet_name}' was not found in the Excel file. Available sheets: {list(worksheet_paths)}"
        )
    return worksheet_path


def read_excel_rows(file_path: str) -> list[list[str]]:
    if file_path.lower().endswith(".xls"):
        raise ExcelReadError(
            "Legacy .xls files are not supported. Upload .xlsx or .csv instead."
        )

    worksheet_path = _worksheet_path_for_file(file_path)
    with zipfile.ZipFile(file_path) as archive:
        shared_strings = _shared_strings_from_zip(archive)
        return _worksheet_rows_from_xml(archive, worksheet_path, shared_strings)


def export_excel_to_csv(file_path: str, csv_path: str) -> None:
    """Convert an Excel sheet to CSV for Spark ingestion."""
    rows = read_excel_rows(file_path)
    with open(csv_path, "w", newline="", encoding="utf-8") as handle:
        csv.writer(handle).writerows(rows)


def _first_row_from_worksheet_xml(
    archive: zipfile.ZipFile,
    worksheet_path: str,
    shared_strings: list[str],
) -> list[str]:
    with archive.open(worksheet_path) as handle:
        root = ET.parse(handle).getroot()

    header_row = next(
        (row for row in root.iter() if _xml_local_tag(row) == "row" and row.get("r") == "1"),
        None,
    )
    if header_row is None:
        raise ExcelReadError("Excel worksheet has no header row")

    cells: list[tuple[int, str]] = []
    for cell in header_row:
        if _xml_local_tag(cell) != "c":
            continue
        cell_ref = cell.get("r")
        if not cell_ref:
            continue
        cells.append(
            (
                _column_index_from_cell_ref(cell_ref),
                _cell_value(cell, shared_strings),
            )
        )

    if not cells:
        raise ExcelReadError("Excel header row is empty")

    cells.sort(key=lambda item: item[0])
    return [value for _, value in cells]


def resolve_excel_sheet_name(file_path: str, sheet_names: list[str] | None = None) -> str:
    names = sheet_names or get_excel_sheet_names(file_path)
    if not names:
        raise ExcelReadError(
            "Excel file has no readable worksheets. Re-save the file in Excel or export as CSV."
        )

    configured_sheet = settings.EXCEL_DEFAULT_SHEET.strip()
    if configured_sheet:
        if configured_sheet not in names:
            raise ExcelReadError(
                f"Configured sheet '{configured_sheet}' was not found. Available sheets: {names}"
            )
        return configured_sheet

    return names[0]


def get_excel_sheet_names(file_path: str) -> list[str]:
    names = _sheet_names_from_zip(file_path)
    if names:
        return names

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def get_excel_data_address(file_path: str) -> str:
    sheet_name = resolve_excel_sheet_name(file_path)
    return f"'{sheet_name}'!A1"


def read_excel_columns(file_path: str) -> list[str]:
    """Return header names, preferring openpyxl and falling back to XML parsing."""
    if file_path.lower().endswith(".xls"):
        raise ExcelReadError(
            "Legacy .xls files are not supported. Upload .xlsx or .csv instead."
        )

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        if workbook.sheetnames:
            sheet_name = resolve_excel_sheet_name(file_path, workbook.sheetnames)
            sheet = workbook[sheet_name]
            first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            return [str(cell).strip() if cell is not None else "" for cell in first_row]
    except StopIteration as exc:
        raise ExcelReadError("Excel worksheet has no header row") from exc
    except KeyError as exc:
        logger.debug("openpyxl could not open worksheet %s, falling back to XML", file_path, exc_info=True)
    except ExcelReadError:
        raise
    except Exception:
        logger.debug("openpyxl failed to read columns from %s, falling back to XML", file_path, exc_info=True)
    finally:
        workbook.close()

    return _read_excel_columns_from_xml(file_path)


def _read_excel_columns_from_xml(file_path: str) -> list[str]:
    worksheet_path = _worksheet_path_for_file(file_path)

    with zipfile.ZipFile(file_path) as archive:
        shared_strings = _shared_strings_from_zip(archive)
        return _first_row_from_worksheet_xml(archive, worksheet_path, shared_strings)
